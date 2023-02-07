from openpyxl import load_workbook

month=int(input("Enter the number of days in a month: "))
update_month=month-1

row_start=int(input("Enter the number of row from where data is starting (excluding heading): "))
row_index_count=2



#HEADER CHUNK

# wb=Workbook()
wb1 = load_workbook('ConsolidatedData.xlsx')

#create an active worksheet:
ws1 = wb1.active

headings=['Date','Diesel_Issuance','Vehicle_No.']

if ws1.cell(row=1, column=1).value=='Date':
    pass
else:
    for i in range(3):
        ws1.cell(row=1, column=i+1).value=headings[i]
wb1.save('ConsolidatedData.xlsx')




#sheet names
file_path = "concatenatefile.xlsx"
master_book =load_workbook(file_path, data_only=True)

sheets = master_book.sheetnames
sheets_names=sheets
sheets_count=len(sheets_names)
wb2 = load_workbook(file_path, data_only=True)

for sheet in range(sheets_count):
    print(sheets_names[sheet])
    
    
    
    ws2=wb2[sheets_names[sheet]]




    #Date Data
    is_data=True
    master_row_data=row_start
    total_cells=update_month+master_row_data
    Date=[]
    while is_data:
        
        data=ws2.cell(row=master_row_data, column=2).value
        master_row_data += 1
        Date.append(data)
        if master_row_data==total_cells+2:
            is_data=False

    Date.pop()
    # print(Data)



    # Deisel Data
    is_data=True
    master_row_data=row_start
    Deisel=[]
    while is_data:
        
        data=ws2.cell(row=master_row_data, column=9).value
        master_row_data += 1
        Deisel.append(data)
        if master_row_data==total_cells+2:
            is_data=False

    Deisel.pop()
    # print(Deisel)
    # print(len(Deisel))
    


    #Vehicle Data

    Vehicle=[]

    wb1 = load_workbook('ConsolidatedData.xlsx')
    sheet_data = sheets_names[sheet]
    #create an active worksheet:
    ws1 = wb1.active

    for i in range(len(Date)):
        ws1.cell(row=i+row_index_count, column=1).value=Date[i]

    for i in range(month):
        ws1.cell(row=i+row_index_count, column=3).value=sheet_data
    

    for i in range(len(Deisel)):
        ws1.cell(row=i+row_index_count, column=2).value=Deisel[i]
    wb1.save('ConsolidatedData.xlsx')

    row_index_count=row_index_count+month
    
    print("Sheet written count = ",sheet)

print('Data is written successfully!')